"""
Descarga el GeoJSON de estados de Mexico
Correr: python scripts/00_descargar_geo.py
"""
import urllib.request
import json
from pathlib import Path

Path('datos/geo').mkdir(parents=True, exist_ok=True)

url = 'https://raw.githubusercontent.com/angelnmara/geojson/master/mexicoHigh.json'

print('Descargando GeoJSON de Mexico...')
urllib.request.urlretrieve(url, 'datos/geo/mexico_estados.json')

with open('datos/geo/mexico_estados.json', encoding='utf-8') as f:
    geo = json.load(f)

print(f'Descargado correctamente')
print(f'   Estados encontrados: {len(geo["features"])}')
print(f'   Propiedades: {geo["features"][0]["properties"]}')
# Diccionario para hacer coincidir nombres
EQUIV = {
    'Coahuila de Zaragoza':              'Coahuila',
    'Michoacán de Ocampo':               'Michoacán',
    'Veracruz de Ignacio de la Llave':   'Veracruz',
}

print('\nVerificando equivalencias...')
import pandas as pd
from openpyxl import load_workbook

wb = load_workbook('datos/ods1/Concentrado_indicadores_de_pobreza_2020.xlsx', read_only=True)
ws = wb['Concentrado estatal']
filas = list(ws.iter_rows(max_row=70, values_only=True))

nombres = [f['properties']['name'] for f in geo['features']]
nombres_df = []
for r in filas[7:]:
    if r[1] and str(r[1]).strip().isdigit():
        nombre = str(r[2])
        nombre_geo = EQUIV.get(nombre, nombre)
        match = '✅' if nombre_geo in nombres else '❌'
        nombres_df.append(nombre_geo)
        print(f'{match} {nombre} -> {nombre_geo}')

print(f'\nTotal coincidencias: {sum(1 for n in nombres_df if n in nombres)}/32')