"""
EmprendeIA_ods · HackODS UNAM 2026
Script 01 — Limpieza ODS 1 (Pobreza)
Correr desde la raíz: python scripts/01_limpieza_ods1_pobreza.py
"""
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

D1      = Path("datos/ods1")
OUT_DIR = D1 / "procesados"
OUT_DIR.mkdir(exist_ok=True)

# ── 1. Pobreza por estado 2010-2020 ──────────────────────────
print("📂 Cargando Concentrado_indicadores_de_pobreza_2020.xlsx ...")
wb = load_workbook(D1 / "Concentrado_indicadores_de_pobreza_2020.xlsx", read_only=True)
ws = wb["Concentrado estatal"]
rows = list(ws.iter_rows(max_row=70, values_only=True))

estados = []
for r in rows[7:]:
    if r[1] and str(r[1]).strip().isdigit():
        estados.append({
            "clave":     str(r[1]).zfill(2),
            "estado":    str(r[2]),
            "pct_2010":  round(float(r[6]),  2) if r[6]  else None,
            "pct_2015":  round(float(r[7]),  2) if r[7]  else None,
            "pct_2020":  round(float(r[8]),  2) if r[8]  else None,
            "pers_2010": int(r[9])  if r[9]  else None,
            "pers_2015": int(r[10]) if r[10] else None,
            "pers_2020": int(r[11]) if r[11] else None,
        })

df_pob = pd.DataFrame(estados).dropna(subset=["pct_2020"])
df_pob.to_csv(OUT_DIR / "pobreza_estados.csv", index=False, encoding="utf-8")
print(f"  ✅ {len(df_pob)} estados → datos/ods1/procesados/pobreza_estados.csv")

# ── 2. Líneas de pobreza 1992-2025 ───────────────────────────
print("📂 Cargando Lineas_de_Pobreza_por_Ingresos_mar2025.xlsx ...")
wb2 = load_workbook(D1 / "Líneas_de_Pobreza_por_Ingresos_mar2025.xlsx", read_only=True)
ws2 = wb2["Líneas de pobreza por ingresos"]
rows2 = list(ws2.iter_rows(max_row=500, values_only=True))

lineas = []
for r in rows2:
    if r[3] and isinstance(r[3], (int, float)) and r[4]:
        try:
            lineas.append({
                "año":            int(r[3]),
                "mes":            str(r[4]),
                "extrema_rural":  round(float(r[6]),  2) if r[6]  else None,
                "extrema_urbano": round(float(r[7]),  2) if r[7]  else None,
                "pobreza_rural":  round(float(r[9]),  2) if r[9]  else None,
                "pobreza_urbano": round(float(r[10]), 2) if r[10] else None,
            })
        except (TypeError, ValueError):
            continue

df_lin = pd.DataFrame(lineas).dropna()
df_lin.to_csv(OUT_DIR / "lineas_pobreza.csv", index=False, encoding="utf-8")
print(f"  ✅ {len(df_lin)} registros → datos/ods1/procesados/lineas_pobreza.csv")

# ── 3. Umbral internacional por estado ───────────────────────
print("📂 Cargando 1.1.1.a_sh_es.csv ...")
df_sh = pd.read_csv(D1 / "1.1.1.a_sh_es.csv", usecols=[0,1,2,3], na_values=["NAN","ND"])
df_sh.columns = ["serie","cvegeo","estado","pct_umbral"]
df_sh = df_sh[df_sh["cvegeo"] != 0].dropna()
df_sh["pct_umbral"] = (
    df_sh["pct_umbral"].astype(str)
    .str.extract(r"([\d.]+)")[0]
    .astype(float).round(2)
)
df_sh.to_csv(OUT_DIR / "umbral_internacional.csv", index=False, encoding="utf-8")
print(f"  ✅ {len(df_sh)} estados → datos/ods1/procesados/umbral_internacional.csv")

print("\n🎉 ODS 1 completado")
print(f"   Pobreza promedio 2020: {df_pob['pct_2020'].mean():.1f}%")
print(f"   Más afectado:  {df_pob.loc[df_pob['pct_2020'].idxmax(),'estado']} ({df_pob['pct_2020'].max():.1f}%)")
print(f"   Menos afectado:{df_pob.loc[df_pob['pct_2020'].idxmin(),'estado']} ({df_pob['pct_2020'].min():.1f}%)")
