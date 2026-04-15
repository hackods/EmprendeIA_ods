"""
EmprendeIA_ods · HackODS UNAM 2026
Script 02 — Limpieza ODS 4 (Educación)
Correr desde la raíz: python scripts/02_limpieza_ods4_educacion.py
"""
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

D4      = Path("datos/ods4")
OUT_DIR = D4 / "procesados"
OUT_DIR.mkdir(exist_ok=True)

# ── 1. Abandono escolar 2000-2024 ─────────────────────────────
print("📂 Cargando Educacion_11.xlsx ...")
wb = load_workbook(D4 / "Educacion_11.xlsx", read_only=True)
ws = wb.active
rows = list(ws.iter_rows(max_row=300, values_only=True))

NIVELES = ["Primaria","Secundaria","Media superior","Superior"]
abandono = []
for r in rows[6:]:
    if r[0] and r[1] and str(r[1]).strip() in NIVELES:
        abandono.append({
            "estado": str(r[0]).strip(),
            "nivel":  str(r[1]).strip(),
            "a2000": r[2], "a2005": r[3], "a2010": r[4],
            "a2015": r[5], "a2020": r[6], "a2022": r[7],
            "a2023": r[8], "a2024": r[9],
        })

df_aban = pd.DataFrame(abandono)
for c in ["a2000","a2005","a2010","a2015","a2020","a2022","a2023","a2024"]:
    df_aban[c] = pd.to_numeric(df_aban[c], errors="coerce")
df_aban.to_csv(OUT_DIR / "abandono_escolar.csv", index=False, encoding="utf-8")
print(f"  ✅ {len(df_aban)} registros → datos/ods4/procesados/abandono_escolar.csv")

# ── 2. Escolaridad promedio 2010-2020 ─────────────────────────
print("📂 Cargando Educacion_05.xlsx ...")
wb2 = load_workbook(D4 / "Educacion_05.xlsx", read_only=True)
ws2 = wb2.active
rows2 = list(ws2.iter_rows(max_row=60, values_only=True))

escol = []
for r in rows2[7:]:
    if r[0] and str(r[0]).strip():
        escol.append({
            "estado":      str(r[0]).strip(),
            "total_2010":  r[1], "hombres_2010": r[2], "mujeres_2010": r[3],
            "total_2015":  r[4], "hombres_2015": r[5], "mujeres_2015": r[6],
            "total_2020":  r[7], "hombres_2020": r[8], "mujeres_2020": r[9],
        })

df_esc = pd.DataFrame(escol).dropna(subset=["total_2020"])
for c in df_esc.columns[1:]:
    df_esc[c] = pd.to_numeric(df_esc[c], errors="coerce").round(2)
df_esc.to_csv(OUT_DIR / "escolaridad_promedio.csv", index=False, encoding="utf-8")
print(f"  ✅ {len(df_esc)} estados → datos/ods4/procesados/escolaridad_promedio.csv")

# ── 3. Índice de finalización por estado ──────────────────────
print("📂 Cargando 4.1.2_dc_1823_es.csv ...")
df_ind = pd.read_csv(D4 / "4.1.2_dc_1823_es.csv", na_values=["NAN","ND","NS"])
col_p = [c for c in df_ind.columns if "primaria" in c.lower() and "Total" in c][0]
col_s = [c for c in df_ind.columns if "secundaria" in c.lower() and "Total" in c][0]
col_m = [c for c in df_ind.columns if "preparatoria" in c.lower() and "Total" in c][0]
df_fin = df_ind[["cvegeo","Entidad_federativa","Periodo", col_p, col_s, col_m]].copy()
df_fin.columns = ["cvegeo","estado","año","idx_primaria","idx_secundaria","idx_preparatoria"]
df_fin = df_fin[df_fin["cvegeo"] != 0].dropna()
for c in ["idx_primaria","idx_secundaria","idx_preparatoria"]:
    df_fin[c] = pd.to_numeric(df_fin[c], errors="coerce").round(2)
df_fin.to_csv(OUT_DIR / "indice_finalizacion.csv", index=False, encoding="utf-8")
print(f"  ✅ {len(df_fin)} registros → datos/ods4/procesados/indice_finalizacion.csv")

print("\n🎉 ODS 4 completado")
nac_ms = df_aban[(df_aban["estado"]=="Estados Unidos Mexicanos") & (df_aban["nivel"]=="Media superior")]
print(f"   Abandono preparatoria 2024: {float(nac_ms['a2024'].values[0]):.1f}%")
nac_esc = df_esc[df_esc["estado"]=="Estados Unidos Mexicanos"]
print(f"   Escolaridad promedio 2020:  {float(nac_esc['total_2020'].values[0]):.2f} años")
